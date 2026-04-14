"""
Local file export for scraped data.
Supports CSV, JSON, and Excel formats with append or create new options.
Allows custom save locations (Desktop, Documents, or custom path).
"""
import csv
import json
import os
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Default export directory
DEFAULT_EXPORT_DIR = Path(__file__).parent.parent / "exports"

# Common save locations (desktop/documents only added if they exist on this OS)
SAVE_LOCATIONS = {
    "default": DEFAULT_EXPORT_DIR,
}
_desktop = Path.home() / "Desktop"
_documents = Path.home() / "Documents"
if _desktop.exists():
    SAVE_LOCATIONS["desktop"] = _desktop
if _documents.exists():
    SAVE_LOCATIONS["documents"] = _documents


def get_save_path(location: str = "default", custom_path: str = None) -> Path:
    """
    Get the save path based on location setting.
    
    Args:
        location: "default", "desktop", "documents", or "custom"
        custom_path: Full path when location is "custom"
    
    Returns:
        Path object for the save directory
    """
    if location == "custom" and custom_path:
        path = Path(custom_path)
    else:
        path = SAVE_LOCATIONS.get(location, DEFAULT_EXPORT_DIR)
    
    path.mkdir(parents=True, exist_ok=True)
    return path


def ensure_export_dir(location: str = "default", custom_path: str = None) -> Path:
    """Ensure the exports directory exists."""
    return get_save_path(location, custom_path)


def get_export_files(location: str = "default", custom_path: str = None):
    """Get list of existing export files from specified location."""
    export_dir = ensure_export_dir(location, custom_path)
    files = []
    for f in export_dir.glob("*"):
        if f.suffix in [".csv", ".json", ".xlsx"]:
            stat = f.stat()
            files.append({
                "name": f.name,
                "path": str(f),
                "size": stat.st_size,
                "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "format": f.suffix[1:]  # Remove the dot
            })
    return sorted(files, key=lambda x: x["modified"], reverse=True)


def export_to_csv(rows: list, filename: str = None, append: bool = False, 
                  location: str = "default", custom_path: str = None) -> dict:
    """
    Export rows to CSV file.
    
    Args:
        rows: List of dictionaries to export
        filename: Optional filename (auto-generated if not provided)
        append: If True, append to existing file; if False, create new
        location: Save location ("default", "desktop", "documents", "custom")
        custom_path: Custom path when location is "custom"
    
    Returns:
        dict with success status, filename, and row count
    """
    if not rows:
        return {"success": False, "error": "No data to export"}
    
    export_dir = ensure_export_dir(location, custom_path)
    
    # Generate filename if not provided
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"scrape_{timestamp}.csv"
    
    # Ensure .csv extension
    if not filename.endswith(".csv"):
        filename += ".csv"
    
    filepath = export_dir / filename
    file_exists = filepath.exists()
    
    # Get column headers from first row
    columns = list(rows[0].keys())
    
    mode = "a" if append and file_exists else "w"
    write_header = not (append and file_exists)
    
    try:
        with open(filepath, mode, newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            
            if write_header:
                writer.writeheader()
            
            writer.writerows(rows)
        
        return {
            "success": True,
            "filename": filename,
            "filepath": str(filepath),
            "rows": len(rows),
            "mode": "appended" if append and file_exists else "created"
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def export_to_json(rows: list, filename: str = None, append: bool = False,
                   location: str = "default", custom_path: str = None) -> dict:
    """
    Export rows to JSON file.
    
    Args:
        rows: List of dictionaries to export
        filename: Optional filename (auto-generated if not provided)
        append: If True, append to existing file; if False, create new
        location: Save location ("default", "desktop", "documents", "custom")
        custom_path: Custom path when location is "custom"
    
    Returns:
        dict with success status, filename, and row count
    """
    if not rows:
        return {"success": False, "error": "No data to export"}
    
    export_dir = ensure_export_dir(location, custom_path)
    
    # Generate filename if not provided
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"scrape_{timestamp}.json"
    
    # Ensure .json extension
    if not filename.endswith(".json"):
        filename += ".json"
    
    filepath = export_dir / filename
    file_exists = filepath.exists()
    
    try:
        if append and file_exists:
            # Load existing data and append
            with open(filepath, "r", encoding="utf-8") as f:
                existing_data = json.load(f)
            
            if isinstance(existing_data, list):
                existing_data.extend(rows)
            else:
                existing_data = rows
            
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(existing_data, f, indent=2, ensure_ascii=False)
        else:
            # Create new file
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(rows, f, indent=2, ensure_ascii=False)
        
        return {
            "success": True,
            "filename": filename,
            "filepath": str(filepath),
            "rows": len(rows),
            "mode": "appended" if append and file_exists else "created"
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def export_to_excel(rows: list, filename: str = None, append: bool = False,
                    location: str = "default", custom_path: str = None) -> dict:
    """
    Export rows to Excel file (.xlsx).
    
    Args:
        rows: List of dictionaries to export
        filename: Optional filename (auto-generated if not provided)
        append: If True, append to existing file; if False, create new
        location: Save location ("default", "desktop", "documents", "custom")
        custom_path: Custom path when location is "custom"
    
    Returns:
        dict with success status, filename, and row count
    """
    if not EXCEL_AVAILABLE:
        return {"success": False, "error": "openpyxl not installed. Run: pip install openpyxl"}
    
    if not rows:
        return {"success": False, "error": "No data to export"}
    
    export_dir = ensure_export_dir(location, custom_path)
    
    # Generate filename if not provided
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"scrape_{timestamp}.xlsx"
    
    # Ensure .xlsx extension
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    
    filepath = export_dir / filename
    file_exists = filepath.exists()
    
    # Get column headers from first row
    columns = list(rows[0].keys())
    
    try:
        if append and file_exists:
            # Load existing workbook and append
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Append rows
            for row in rows:
                ws.append([row.get(col, "") for col in columns])
        else:
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Scraped Data"
            
            # Write header
            ws.append(columns)
            
            # Write data rows
            for row in rows:
                ws.append([row.get(col, "") for col in columns])
            
            # Auto-adjust column widths (approximate)
            for i, col in enumerate(columns, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(50, max(12, len(col) + 2))
        
        wb.save(filepath)
        
        return {
            "success": True,
            "filename": filename,
            "filepath": str(filepath),
            "rows": len(rows),
            "mode": "appended" if append and file_exists else "created"
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def export_data(rows: list, format: str = "csv", filename: str = None, append: bool = False,
                location: str = "default", custom_path: str = None) -> dict:
    """
    Export data to local file.
    
    Args:
        rows: List of dictionaries to export
        format: "csv", "json", or "xlsx"
        filename: Optional filename
        append: If True, append to existing file
        location: Save location ("default", "desktop", "documents", "custom")
        custom_path: Custom path when location is "custom"
    
    Returns:
        dict with export result
    """
    if format == "json":
        return export_to_json(rows, filename, append, location, custom_path)
    elif format == "xlsx":
        return export_to_excel(rows, filename, append, location, custom_path)
    else:
        return export_to_csv(rows, filename, append, location, custom_path)


def get_available_formats():
    """Get list of available export formats."""
    formats = ["csv", "json"]
    if EXCEL_AVAILABLE:
        formats.append("xlsx")
    return formats
